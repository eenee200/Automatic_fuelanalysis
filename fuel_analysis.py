import json
from datetime import datetime, timedelta
import tempfile
from bs4 import BeautifulSoup
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment,Font,PatternFill

# def parse_data(raw_data):
#     data_str = raw_data.strip()
#     data_points = [point.strip('[]').split(',') for point in data_str.split('],[')]
    
#     # First pass to collect raw data
#     raw_parsed_data = []
#     for point in data_points:
#         timestamp = float(point[0])
#         fuel = float(point[1].strip('"'))
#         raw_parsed_data.append((timestamp, fuel))
    
#     # Check if all fuel values are the same
#     fuel_values = set(fuel for _, fuel in raw_parsed_data)
#     if len(fuel_values) == 1:
#         return None
#     if fuel > 300:
#         return None
    
#     # Initialize parameters
#     window_size = 6  # Size of the sliding window
#     look_back_window = 6  # Number of previous points to check
#     parsed_data = []
#     last_valid_fuel = None
    
#     def validate_fuel_level(current_idx, current_value):
#         """
#         Validate fuel level by checking previous values
#         Returns the appropriate fuel value to use
#         """
#         if current_value <= 3:
#             return last_valid_fuel if last_valid_fuel is not None else current_value
            
#         # Look back to validate sudden drops
#         if last_valid_fuel is not None and current_value < last_valid_fuel * 0.5:
#             previous_values = []
#             start_idx = max(0, current_idx - look_back_window)
            
#             for j in range(start_idx, current_idx):
#                 prev_value = raw_parsed_data[j][1]
#                 if prev_value > 3:
#                     previous_values.append(prev_value)
            
#             if previous_values:
#                 # Use median of previous values if they're closer to last_valid_fuel
#                 median_previous = sorted(previous_values)[len(previous_values)//2]
#                 if abs(median_previous - last_valid_fuel) < abs(current_value - last_valid_fuel):
#                     return median_previous
        
#         return current_value
    
#     # Process data with running median
#     for i in range(len(raw_parsed_data)):
#         timestamp = raw_parsed_data[i][0]
#         current_value = raw_parsed_data[i][1]
        
#         # Calculate window boundaries
#         start_idx = max(0, i - window_size + 1)
#         window_values = []
        
#         # Process each value in the window
#         for j in range(start_idx, i + 1):
#             window_value = raw_parsed_data[j][1]
#             validated_value = validate_fuel_level(j, window_value)
#             if validated_value > 3:
#                 window_values.append(validated_value)
        
#         # Calculate median for non-empty windows
#         if window_values:
#             median_fuel = sorted(window_values)[len(window_values)//2]
#             # Validate the median against the last valid fuel
#             if last_valid_fuel is not None:
#                 median_fuel = validate_fuel_level(i, median_fuel)
#             if median_fuel > 3:
#                 last_valid_fuel = median_fuel
#         else:
#             median_fuel = last_valid_fuel if last_valid_fuel is not None else current_value
        
#         parsed_data.append((timestamp, median_fuel))
    
#     # Final check for all values being the same
#     final_fuel_values = set(fuel for _, fuel in parsed_data)
#     if len(final_fuel_values) == 1:
#         return None
    
#     return parsed_data
def parse_data(raw_data):
    data_str = raw_data.strip()
    data_points = [point.strip('[]').split(',') for point in data_str.split('],[')]
    valid_fuel = None
    
    parsed_data = []
    fuel_values = set()
    for point in data_points:
        timestamp = float(point[0])
        fuel = float(point[1].strip('"'))
        if fuel != 0:
            valid_fuel = fuel
        if fuel == 0 and valid_fuel is not None:   
            fuel = valid_fuel
        parsed_data.append((timestamp, fuel))
        fuel_values.add(fuel)
    
    # Check if all fuel values are the same
    if len(fuel_values) == 1:
        return None  # Return None if all fuel values are the same
    # if fuel > 700:
    #     return None
    else:
    
        return parsed_data


def load_data_from_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    datasets = []
    identifiers = []
    date_range = None
    identifiers_to_remove = set()

    # Find all script tags and object rows
    all_elements = soup.find_all(['script', 'table'])
    current_identifier = None
    current_datasets = []

    for element in all_elements:
        if element.name == 'table':
            # Check for object identifier
            object_row = element.find('td', string='Обьект:')
            if object_row:
                # Process previous identifier's datasets if exists
                if current_identifier and current_datasets:
                    valid_datasets = []
                    for dataset in current_datasets:
                        parsed_data = parse_data(dataset)
                        if parsed_data is not None:
                            valid_datasets.append(dataset)
                            break  # Only keep the first valid dataset
                    
                    # Add to identifiers and datasets, or mark for removal
                    if valid_datasets:
                        datasets.extend(valid_datasets)  # Will only contain the first valid dataset
                        identifiers.append(current_identifier)
                    else:
                        identifiers_to_remove.add(current_identifier)

                # Reset for new identifier
                current_identifier = element.find('td', string='Обьект:').find_next_sibling('td').get_text(strip=True)
                current_datasets = []

                # Check for date range
                date_row = element.find('td', string='Хугацаа:')
                if date_row:
                    date_cell = date_row.find_next_sibling('td')
                    if date_cell and not date_range:
                        date_range = date_cell.get_text(strip=True)

        elif element.name == 'script' and element.string:
            # Collect fuel data matches
            data_matches = re.findall(r'data":\s*(\[.*?\])\s*,\s*"data_index', element.string, re.DOTALL)
            current_datasets.extend(data_matches)

    # Process last identifier's datasets
    if current_identifier and current_datasets:
        valid_datasets = []
        for dataset in current_datasets:
            parsed_data = parse_data(dataset)
            if parsed_data is not None:
                valid_datasets.append(dataset)
                break  # Only keep the first valid dataset
        
        if valid_datasets:
            datasets.extend(valid_datasets)  # Will only contain the first valid dataset
            identifiers.append(current_identifier)
        else:
            identifiers_to_remove.add(current_identifier)

    # Remove identifiers with no valid data
    final_identifiers = [ident for ident in identifiers if ident not in identifiers_to_remove]
    
    if not datasets:
        raise ValueError("No data arrays found in the file.")

    return datasets, final_identifiers, [date_range] if date_range else ['']
def load_daily_distances(file_path, valid_identifiers):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    all_daily_distances = []
    all_daily_dates = []
    found_identifiers = []
    deleted_distances = []
    deleted_dates = []
    deleted_identifiers = []

    object_rows = soup.find_all('td', string='Обьект:')
    
    for object_row in object_rows:
        identifier = object_row.find_next_sibling('td').get_text(strip=True)
        current_table = object_row.find_parent('table')
        distance_table = current_table.find_next_sibling('table')
        
        if distance_table:
            daily_distances = []
            daily_dates = []
            for row in distance_table.find_all('tr')[1:]:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    date = datetime.strptime(cells[0].get_text(strip=True), '%Y-%m-%d').date()
                    distance = cells[1].get_text(strip=True)
                    distance_val = float(distance.replace(' km', ''))
                    daily_distances.append(distance_val)
                    daily_dates.append(date)
            
            if identifier in valid_identifiers:
                found_identifiers.append(identifier)
                all_daily_distances.append(daily_distances)
                all_daily_dates.append(daily_dates)
            else:
                deleted_identifiers.append(identifier)
                deleted_distances.append(daily_distances)
                deleted_dates.append(daily_dates)

    ordered_distances = []
    ordered_dates = []
    for identifier in valid_identifiers:
        if identifier in found_identifiers:
            idx = found_identifiers.index(identifier)
            ordered_distances.append(all_daily_distances[idx])
            ordered_dates.append(all_daily_dates[idx])

    final_distances = ordered_distances + deleted_distances
    final_dates = ordered_dates + deleted_dates
    final_identifiers = valid_identifiers + deleted_identifiers

    return final_distances, final_identifiers, final_dates






def detect_refills(data, threshold_percentage=5, time_window_minutes=60):
    refills = []
    in_refill = False
    min_fuel = None
    max_fuel = None
    start_time = None
    last_valid_fuel = None  # To store the last fuel value greater than 3

    def check_previous_fuel_levels(data, current_index, start_time, max_fuel):
        """Check if there's a higher fuel level in the previous 10 minutes"""
        check_start_time = start_time - timedelta(minutes=120)
        comparison_fuel = max_fuel - 5
        check_end_time = start_time
        for j in range(current_index - 1, -1, -1):
            check_time = datetime.utcfromtimestamp(data[j][0]/1000)
            if check_time < check_start_time:
                break
            if check_time > check_end_time:
                continue
            if data[j][1] > comparison_fuel:
                return True
        return False
    
    def find_real_start_time(data, start_idx, start_fuel):
        """Find the actual start time by skipping over periods of constant fuel level"""
        real_start_idx = start_idx
        current_fuel = start_fuel
        
        for i in range(start_idx + 1, len(data)):
            if data[i][1] > current_fuel:
                real_start_idx = i - 1
                break
            if data[i][1] < current_fuel:
                break
        
        return datetime.utcfromtimestamp(data[real_start_idx][0]/1000)
    
    for i in range(1, len(data)):
        prev_fuel = data[i-1][1]
        current_fuel = data[i][1]
        current_time = datetime.utcfromtimestamp(data[i][0]/1000)

        if current_fuel== None:
            current_fuel = 0
        if prev_fuel== None:
            prev_fuel = 0
        
        if current_fuel >= 1:
            last_valid_fuel = current_fuel
        
        if current_fuel >= prev_fuel:
            if not in_refill:
                in_refill = True
                min_fuel = prev_fuel if prev_fuel >= 1 else last_valid_fuel
                start_time = find_real_start_time(data, i-1, prev_fuel)
            max_fuel = current_fuel
            last_time = current_time
        elif in_refill:
            in_refill = False
            if min_fuel is not None and max_fuel is not None:
                if min_fuel <= 0 and last_valid_fuel is not None:
                    min_fuel = last_valid_fuel
                
                percent_change = max_fuel - min_fuel
                if min_fuel >= 0:
                    if percent_change > threshold_percentage:
                        valid_refill = True
                        end_time = last_time + timedelta(minutes=time_window_minutes)
                        
                        if check_previous_fuel_levels(data, i, start_time, max_fuel):
                            valid_refill = False
                        else:
                            # Check for significant drops after the refill
                            for j in range(i, len(data)):
                                check_time = datetime.utcfromtimestamp(data[j][0]/1000)
                                if check_time > end_time:
                                    break
                                # Only invalidate if we see a significant drop
                                if data[j][1] <= min_fuel + (percent_change * 0.5):  # Allow for some normal usage drop
                                    valid_refill = False
                                    break
                        
                        if valid_refill:
                            if refills and (last_time - refills[-1]['timestamp']) <= timedelta(minutes=time_window_minutes):
                                refills[-1]['max_fuel'] = max(refills[-1]['max_fuel'], max_fuel)
                                refills[-1]['percent_change'] = refills[-1]['max_fuel'] - refills[-1]['min_fuel']
                            else:
                                refills.append({
                                    'timestamp': start_time,
                                    'percent_change': percent_change,
                                    'max_fuel': max_fuel,
                                    'min_fuel': min_fuel
                                })
            min_fuel, max_fuel = None, None
    
    return refills
def analyze_fuel_data(raw_data):
    data = parse_data(raw_data)
    
    if not data:
        return [], {'num_refills': 0, 'first_fuel': None, 'last_fuel': None}
    
    first_fuel = data[0][1]
    last_fuel = data[-1][1]
    refills = detect_refills(data)
    
    stats = {
        'num_refills': len(refills),
        'first_fuel': first_fuel,
        'last_fuel': last_fuel,
    }
    
    return refills, stats

def export_to_excel(datasets, identifiers, date_ranges, all_daily_distances, all_daily_dates, output_file='fuel_analysis.xlsx'):
    try:
        all_summary_data = []
        all_refills_data = []
        all_daily_data = []
        urgent_check_needed = []
        
        # Process multiple datasets
        for idx, (refills, stats, data) in enumerate(datasets):
            dataset_name = identifiers[idx] 
            refills_data = []
            total_refill = 0.0  # Initialize as float
            total_consumption = 0.0  # Initialize as float
            
            # Safely handle None values for first and last fuel readings
            first = float(stats['first_fuel'] if stats['first_fuel'] is not None else 0)
            last = float(stats['last_fuel'] if stats['last_fuel'] is not None else 0)
            daily_distances = all_daily_distances[idx] if idx < len(all_daily_distances) else []

            # Create a dictionary to store refill dates for daily counting
            refill_dates = {}
            for refill in refills:
                refill_date = refill['timestamp'].date()
                refill_dates[refill_date] = refill_dates.get(refill_date, 0) + 1
            
            for i, refill in enumerate(refills, 1):
                # Safely handle None values in refill calculations
                min_fuel = float(refill.get('min_fuel', 0) or 0)  # Convert None to 0
                max_fuel = float(refill.get('max_fuel', 0) or 0)  # Convert None to 0
                
                consumption = round(first - min_fuel, 2)
                percent_change = max_fuel - min_fuel
                
                total_refill += percent_change
                total_consumption += consumption

                refills_data.append({
                    ' ': " ",
                    'Эхэлсэн хугацаа': refill['timestamp'],
                    'Өмнөх түлш': round(min_fuel, 2),
                    'Дараах түлш': round(max_fuel, 2),
                    'Нэмсэн түлш': round(percent_change, 2),
                    'Сүүлд дүүргэснээс хойш зарцуулалт': round(consumption, 2)
                })

                first = max_fuel

            # Safely calculate final consumption
            final_consumption = first - last if first is not None and last is not None else 0
            total_consumption += final_consumption

            # Calculate distance per refill safely
            total_distance = sum(daily_distances) if daily_distances else 0
            avg_consumption = (total_consumption / total_distance * 100) if total_distance > 0 else 0

            # Create summary data with safe handling of None values
            summary_data = {
                'Обьект': dataset_name + (" (яаралтай шалгуулах хэрэгтэй)" if idx in urgent_check_needed else ""),
                'Нийт явсан км': total_distance if total_distance != 0 else 'N/A',
                'Түлш дүүрлт /Л/': round(float(total_refill), 2),
                'Түлш дүүргэсэн тоо': int(stats.get('num_refills', 0)),
                'Түлш зарцуулалт /Л/': round(float(total_consumption), 2),
                'Дундаж хэрэглээ/100км/': round(avg_consumption, 2) if avg_consumption > 0 else "",
                'Эхний үлдэгдэл': round(float(stats.get('first_fuel', 0) or 0), 2),
                'Эцсийн үлдэгдэл': round(float(stats.get('last_fuel', 0) or 0), 2)
            }
            all_summary_data.append(summary_data)

            # Check for urgent cases
            if stats.get('first_fuel', 0) == 0 and stats.get('last_fuel', 0) == 0:
                urgent_check_needed.append(idx)

            # Check for multiple refills in 24 hours
            refill_times = [refill['timestamp'] for refill in refills]
            refill_times.sort()
            for i in range(len(refill_times)):
                end_time = refill_times[i]
                start_time = end_time - timedelta(hours=24)
                refills_in_24h = sum(1 for t in refill_times if start_time <= t <= end_time)
                if refills_in_24h >= 5:
                    urgent_check_needed.append(idx)
                    break

            # Process daily data
            daily_data = []
            daily_dates = all_daily_dates[idx] if idx < len(all_daily_dates) else []
            daily_start_fuel = 0.0
            daily_end_fuel = 0.0
            
            for date_idx, current_date in enumerate(daily_dates):
                # Find fuel levels for this date
                day_fuel_levels = []
                total_daily_refill = 0.0
                
                for timestamp, fuel_level in data:
                    data_date = datetime.utcfromtimestamp(timestamp / 1000).date()
                    if data_date == current_date:
                        day_fuel_levels.append(float(fuel_level if fuel_level is not None else 0))
                        
                # Get start and end fuel levels for the day
                if day_fuel_levels:
                    daily_start_fuel = day_fuel_levels[0]
                    daily_end_fuel = day_fuel_levels[-1]
                
                # Calculate total daily refill amount
                for refill in refills:
                    if refill['timestamp'].date() == current_date:
                        total_daily_refill += float(refill.get('percent_change', 0) or 0)
                
                daily_consumption = daily_start_fuel + total_daily_refill - daily_end_fuel
                daily_distance = float(daily_distances[date_idx] if date_idx < len(daily_distances) else 0)
                
                # Calculate average consumption per 100km
                avg_consumption = (daily_consumption / daily_distance * 100) if daily_distance > 0 else 0
                
                daily_data.append({
                    '': current_date,
                    'Нийт явсан км': round(daily_distance, 2),
                    'Түлш дүүрлт /Л/': round(total_daily_refill, 2),
                    'Түлш дүүргэсэн тоо': refill_dates.get(current_date, 0),
                    'Түлш зарцуулалт /Л/': round(daily_consumption, 2),
                    'Дундаж хэрэглээ/100км/': round(avg_consumption, 2) if avg_consumption > 0 else " ",
                    'Эхний үлдэгдэл': round(daily_start_fuel, 2),
                    'Эцсийн үлдэгдэл': round(daily_end_fuel, 2),
                })
            
            all_refills_data.append((dataset_name, refills_data))
            all_daily_data.append((dataset_name, daily_data))

        # Export all data to Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Ерөнхий мэдээлэл'
        
        date_range_cell = worksheet.cell(row=1, column=1, value=f"Хугацаа: {date_ranges[0]}")
        date_range_cell.font = Font(bold=True)

        # Write summary data
        summary_df = pd.DataFrame(all_summary_data)
        header_row = dataframe_to_rows(summary_df, index=False, header=True)
        worksheet.append(next(header_row))
        
       
        # Add datasets to the Excel file
        for idx, (dataset_name, refills_data) in enumerate(all_refills_data):
               # Write object name (identifier) and date range only once at the top
            # Write summary for each dataset
            daily_distances = all_daily_distances[idx] if idx < len(all_daily_distances) else []
            summary_start_row = worksheet.max_row   # Leave a gap before the next dataset
            total_distance = sum(daily_distances) if daily_distances else 0
            total_refill = all_summary_data[idx]['Түлш зарцуулалт /Л/']

    # Calculate distance per refill, handle division by zero if total_refill is 0
            distance_per_refill = total_refill / total_distance if total_distance != 0 else 'N/A'
            distance_per_refill = distance_per_refill * 100
            summary_row = {
                'Обьект': dataset_name + (" (яаралтай шалгуулах хэрэгтэй)" if idx in urgent_check_needed else ""),
                'Нийт явсан км': sum(daily_distances) if daily_distances else 'N/A',
                'Түлш дүүрлт /Л/': round(all_summary_data[idx]['Түлш дүүрлт /Л/'], 2),
                'Түлш дүүргэсэн тоо': all_summary_data[idx]['Түлш дүүргэсэн тоо'],
                'Түлш зарцуулалт /Л/': round(all_summary_data[idx]['Түлш зарцуулалт /Л/'], 2),
                'Дундаж хэрэглээ/100км/': round(distance_per_refill, 2) if isinstance(distance_per_refill, (int, float)) else 'N/A',
                'Эхний үлдэгдэл': round(all_summary_data[idx]['Эхний үлдэгдэл'], 2),
                'Эцсийн үлдэгдэл': round(all_summary_data[idx]['Эцсийн үлдэгдэл'], 2)
            }
            worksheet.append([summary_row[key] for key in summary_row])
            if idx in urgent_check_needed:
                cell = worksheet.cell(row=worksheet.max_row, column=1)
                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red background
            else:
                cell = worksheet.cell(row=worksheet.max_row, column=1)
                cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  
            fill_color = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            for col in range(1, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=2, column=col)
                cell.fill = fill_color
            # Write refills data
            refills_df = pd.DataFrame(refills_data)
            daily_row = worksheet.max_row + 1
            for r in dataframe_to_rows(refills_df, index=False, header=True):
                worksheet.append(r)
            fil_color = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
            for col in range(2, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=daily_row, column=col)
                cell.fill = fil_color

            daily_df = pd.DataFrame(all_daily_data[idx][1])
            daily_start_row = worksheet.max_row + 1
            
            for r in dataframe_to_rows(daily_df, index=False, header=True):
                worksheet.append(r)
            
            # Get the starting and ending rows for daily data
            daily_end_row = worksheet.max_row
            worksheet.row_dimensions.group(daily_start_row , daily_end_row, outline_level=1, hidden=True)
            for col in range(2, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=daily_start_row, column=col)
                cell.fill = fil_color
            

        # Apply formatting
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        
        # Apply border and alignment to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_file)

        return output_file, len(datasets)

    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        return None, 0


def main(file_path1, file_path2):
    all_datasets = []
    all_identifiers = []
    all_date_ranges = []

    # Load datasets from the first HTML file
    raw_datasets, active_identifiers, date_ranges = load_data_from_file(file_path1)
    all_identifiers.extend(active_identifiers)
    all_date_ranges.extend(date_ranges)

    # Process each dataset from file_path1
    for raw_data in raw_datasets:
        refills, stats = analyze_fuel_data(raw_data)
        data = parse_data(raw_data)
        all_datasets.append((refills, stats, data))

    # Load all daily distances and dates
    all_daily_distances, combined_identifiers, all_daily_dates = load_daily_distances(file_path2, active_identifiers)
    
    # Find removed identifiers
    removed_identifiers = [id for id in combined_identifiers if id not in active_identifiers]
    all_identifiers = active_identifiers + removed_identifiers

    # Create empty datasets for removed identifiers
    for i in range(len(removed_identifiers)):
        idx = combined_identifiers.index(removed_identifiers[i])
        daily_dates = all_daily_dates[idx]
        
        empty_refills = []
        empty_stats = {'num_refills': 0, 'first_fuel': 0, 'last_fuel': 0}
        empty_data = []

        if daily_dates:
            # Create data points for each day
            for date in daily_dates:
                timestamp = int(datetime.combine(date, datetime.min.time()).timestamp() * 1000)
                empty_data.append((timestamp, 0))
        
        all_datasets.append((empty_refills, empty_stats, empty_data))

    # Create temporary file for Excel output
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_path = tmp.name

    # Export data to Excel
    excel_file, num_datasets = export_to_excel(
        all_datasets, 
        all_identifiers, 
        all_date_ranges, 
        all_daily_distances,
        all_daily_dates,
        output_file=temp_path
    )

    return excel_file, num_datasets

if __name__ == "__main__":
    file_path1 = 'C:/Users/User/Desktop/ttt/eh.html'
    file_path2 = 'C:/Users/User/Desktop/ttt/tog.html'
    excel_file, num_datasets = main(file_path1, file_path2)
    if excel_file:
        print(f"Analysis of {num_datasets} datasets exported to {excel_file}")
    else:
        print("Failed to export analysis to Excel")